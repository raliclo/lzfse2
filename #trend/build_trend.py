"""
Build lzfse2_R1_R45_trend.xlsx
Formats: TGZ, BVX3, Other3, Optimal3, Lazy2, Optimal, Apple, ZSTD, TLZ4 (9 formats)
Metrics: Encode/Decode Speed, RSS, Energy, Compress Ratio
Charts:  Actual LineChart objects embedded in each chart sheet

Era colour coding:
  Yellow  R3-R22  : single-run, speed + ratio only
  Green   R23-R24 : single-run, speed + ratio + RSS
  Blue    R25-R26 : n=40, speed + RSS + ratio
  White   R27-R45 : n=40, full (speed + RSS + energy + ratio)
"""
import subprocess, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel

ROUND_COMMITS = {
    'R3':  '420a171', 'R4':  '1e1191a', 'R5':  'a0bcb3c',
    'R6':  '278ee42', 'R7':  '1e77794', 'R8':  '5920c2d',
    'R9':  'b17575b', 'R10': 'ce43110', 'R11': 'f027fed',
    'R12': '78786d3', 'R13': 'aca95cc', 'R15': '95f2071',
    'R16': 'd7983f5', 'R17': '02789a8', 'R18': 'c5fa99c',
    'R19': 'ac96003', 'R21': '60664ad', 'R22': '6f1db30',
    'R23': '7b87914', 'R24': '2b0bf93', 'R25': '91d356b',
    'R26': 'b9d35b7',
}
ROUND_ORDER = [
    'R3','R4','R5','R6','R7','R8','R9','R10','R11','R12','R13',
    'R15','R16','R17','R18','R19','R21','R22',
    'R23','R24','R25','R26',
    'R27','R28','R29','R30','R31','R32','R33','R34',
    'R35','R36','R37','R38','R39','R40','R41',
    'R42','R43','R44','R45',
]
FORMATS = ['TGZ', 'BVX3', 'Other3', 'Optimal3', 'Lazy2', 'Optimal', 'Apple', 'ZSTD', 'TLZ4']
FORMAT_KEYS = {
    'TGZ':            'TGZ',
    'LZFSE (BVX3)':   'BVX3',
    'LZFSE (Other3)': 'Other3',
    'LZFSE (Optimal3)': 'Optimal3',
    'LZFSE (Lazy2)':  'Lazy2',
    'LZFSE (Optimal)':'Optimal',
    'LZFSE (Apple)':  'Apple',
    'ZSTD':           'ZSTD',
    'TLZ4':           'TLZ4',
}
NF = len(FORMATS)  # 9

# Column layout in Raw sheet (1-indexed), NF=7:
#  A=1   : Round
#  B-H   : Enc Speed   (cols 2-8)
#  I-O   : Dec Speed   (cols 9-15)
#  P-V   : Enc RSS     (cols 16-22)
#  W-AC  : Dec RSS     (cols 23-29)
#  AD-AJ : Enc Energy  (cols 30-36)
#  AK-AQ : Dec Energy  (cols 37-43)
#  AR-AX : Ratio       (cols 44-50)
COL_ENC_SPD  = 2
COL_DEC_SPD  = 2  + NF  # 9
COL_ENC_RSS  = 2  + 2*NF  # 16
COL_DEC_RSS  = 2  + 3*NF  # 23
COL_ENC_ENJ  = 2  + 4*NF  # 30
COL_DEC_ENJ  = 2  + 5*NF  # 37
COL_RATIO    = 2  + 6*NF  # 44

ERA_FILL = {
    'early': PatternFill("solid", fgColor="FFF2CC"),
    'rss':   PatternFill("solid", fgColor="E2EFDA"),
    'n40':   PatternFill("solid", fgColor="DEEAF1"),
    'full':  PatternFill("solid", fgColor="FFFFFF"),
}
SERIES_COLORS = {
    'TGZ':    '7F7F7F',
    'BVX3':   'ED7D31',
    'Other3': '4472C4',
    'Optimal3': '5B9BD5',
    'Lazy2':  '70AD47',
    'Optimal':'FFC000',
    'Apple':  '00B0F0',
    'ZSTD':   'FF0000',
    'TLZ4':   'A020F0',
}

def era_of(rnd):
    n = int(rnd[1:])
    if n <= 22: return 'early'
    if n <= 24: return 'rss'
    if n <= 26: return 'n40'
    return 'full'

def pf(v):
    try: return float(v.strip()) if v and v.strip() else None
    except: return None

import pathlib as _pathlib
SCRIPT_DIR = _pathlib.Path(__file__).parent
REPO_ROOT = str(_pathlib.Path.home() / 'proj' / 'lzfse2')

def git_csv(commit):
    if commit == 'local':
        try:
            with open(f'{REPO_ROOT}/BenchMarkResult.csv', encoding='utf-8-sig') as f:
                return f.read()
        except Exception:
            return ''
    if commit == 'win-local':
        try:
            with open(f'{REPO_ROOT}/helper_windows/bench_results_csv/BenchMarkResult-Win.csv', encoding='utf-8-sig') as f:
                return f.read()
        except Exception:
            return ''
    r = subprocess.run(
        ['git', 'show', f'{commit}:BenchMarkResult.csv'],
        capture_output=True, text=True, errors='replace',
        cwd=REPO_ROOT
    )
    return r.stdout if r.returncode == 0 else ''

def extract_row(commit):
    raw = git_csv(commit)
    if not raw: return {}
    lines = [l for l in raw.splitlines() if l.strip()]
    if not lines: return {}
    header = [h.strip() for h in lines[0].split(',')]
    has_n      = any(h in ('n', 'N') for h in header[:3])
    has_rss    = any('RSS' in h for h in header)
    has_energy = any('energy' in h.lower() for h in header)

    result = {}
    for line in lines[1:]:
        cols = [c.strip() for c in line.split(',')]
        while len(cols) < len(header): cols.append('')

        if has_n:
            dataset, n_val, fmt_raw = cols[0], cols[1], cols[2]
            if n_val != '40': continue
            enc_spd = pf(cols[7]); dec_spd = pf(cols[8]); ratio = pf(cols[9])
            enc_rss = pf(cols[12]) if len(cols) > 12 else None
            dec_rss = pf(cols[13]) if len(cols) > 13 else None
            enc_ej = dec_ej = None
            if has_energy:
                for i, h in enumerate(header):
                    hl = h.lower()
                    if 'encode_cpu_energy_j' in hl: enc_ej = pf(cols[i]) if i < len(cols) else None
                    if 'decode_cpu_energy_j' in hl: dec_ej = pf(cols[i]) if i < len(cols) else None
        elif has_rss:
            dataset, fmt_raw = cols[0], cols[1]
            enc_spd = pf(cols[6]); dec_spd = pf(cols[7]); ratio = pf(cols[8])
            enc_rss = pf(cols[11]) if len(cols) > 11 else None
            dec_rss = pf(cols[12]) if len(cols) > 12 else None
            enc_ej = dec_ej = None
        else:
            dataset, fmt_raw = cols[0], cols[1]
            enc_spd = pf(cols[6]); dec_spd = pf(cols[7]); ratio = pf(cols[8])
            enc_rss = dec_rss = enc_ej = dec_ej = None

        if fmt_raw not in FORMAT_KEYS: continue
        if dataset not in ('claw-code', 'llama.cpp'): continue
        fmt = FORMAT_KEYS[fmt_raw]
        if dataset not in result: result[dataset] = {}
        result[dataset][fmt] = dict(
            enc_spd=enc_spd, dec_spd=dec_spd,
            enc_rss=enc_rss, dec_rss=dec_rss,
            enc_ej=enc_ej,   dec_ej=dec_ej,
            ratio=ratio
        )
    return result

# All R27-R45 commits — extract all formats directly from git/local CSV
R27_45_COMMITS = {
    'R27': 'd6987a0', 'R28': '55da799', 'R29': 'c3b5af8', 'R30': '996e16e',
    'R31': '068737c', 'R32': '2fa46a7', 'R33': '3b38b4d', 'R34': 'c1b510d',
    'R35': 'c2f0721', 'R36': 'edaa07e', 'R37': '7dfe4b3', 'R38': 'd1247d6',
    'R39': '65d1e67', 'R40': 'c288335',
    'R41': '7a46d57',
    'R42': 'd8563cc',
    'R42': '0286a0d',
    'R43': '1cef7e9',
    'R44': 'd5b35ec',
    # R45 is the latest Windows retest round and lives in the Windows CSV.
    'R45': 'win-local',
}
ALL_COMMITS = {**ROUND_COMMITS, **R27_45_COMMITS}

# ── Extract data ─────────────────────────────────────────────────────────
print("Extracting all rounds from git history...")
all_data = {}
for rnd in ROUND_ORDER:
    data = extract_row(ALL_COMMITS[rnd])
    all_data[rnd] = data
    cc = list(all_data[rnd].get('claw-code', {}).keys())
    print(f"  {rnd}: {cc}")

# ── Build xlsx ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

HDR_BLUE  = PatternFill("solid", fgColor="2E75B6")
HDR_FONT  = Font(bold=True)
HDR_FONTW = Font(bold=True, color="FFFFFF")

def make_raw(wb, sname, dataset):
    ws = wb.create_sheet(sname)
    groups = [
        ('Encode Speed (MB/s)', COL_ENC_SPD),
        ('Decode Speed (MB/s)', COL_DEC_SPD),
        ('Encode RSS (MB)',      COL_ENC_RSS),
        ('Decode RSS (MB)',      COL_DEC_RSS),
        ('Encode Energy (J)',    COL_ENC_ENJ),
        ('Decode Energy (J)',    COL_DEC_ENJ),
        ('Compress Ratio',       COL_RATIO),
    ]
    # Row 1: group headers
    ws.cell(1, 1, 'Round').font = HDR_FONTW
    ws.cell(1, 1).fill = HDR_BLUE
    for title, sc in groups:
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc+NF-1)
        c = ws.cell(1, sc, title)
        c.font = HDR_FONTW; c.fill = HDR_BLUE
        c.alignment = Alignment(horizontal='center')
    # Row 2: format names
    for sc in [COL_ENC_SPD, COL_DEC_SPD, COL_ENC_RSS, COL_DEC_RSS,
               COL_ENC_ENJ, COL_DEC_ENJ, COL_RATIO]:
        for i, fmt in enumerate(FORMATS):
            ws.cell(2, sc+i, fmt).font = HDR_FONT
    # Data
    for rnd in ROUND_ORDER:
        d = all_data[rnd].get(dataset, {})
        fill = ERA_FILL[era_of(rnd)]
        row = [rnd]
        for metric in ['enc_spd','dec_spd','enc_rss','dec_rss','enc_ej','dec_ej']:
            for fmt in FORMATS:
                row.append(d.get(fmt, {}).get(metric))
        for fmt in FORMATS:
            row.append(d.get(fmt, {}).get('ratio'))
        ws.append(row)
        ri = ws.max_row
        for col in range(1, COL_RATIO + NF):
            ws.cell(ri, col).fill = fill
        ws.cell(ri, 1).font = HDR_FONT
    ws.column_dimensions['A'].width = 8
    for col in range(2, COL_RATIO + NF):
        ws.column_dimensions[get_column_letter(col)].width = 10
    return ws

def add_line_chart(ws, title, raw_sname, data_min_col, data_max_col,
                   anchor, y_title, categories_ref):
    """Add a LineChart to ws at anchor cell (e.g. 'J3')."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = y_title
    chart.x_axis.title = 'Round'
    chart.height = 14
    chart.width  = 24

    n_rows = len(ROUND_ORDER)
    raw_ws = ws.parent[raw_sname]

    for col_offset in range(data_max_col - data_min_col + 1):
        col = data_min_col + col_offset
        fmt_name = FORMATS[col_offset % NF]
        ref = Reference(raw_ws, min_col=col, max_col=col,
                        min_row=3, max_row=2 + n_rows)
        series = chart.series.append(None)
        chart.append(ref)
        s = chart.series[-1]
        s.title = SeriesLabel(v=fmt_name)
        color = SERIES_COLORS.get(fmt_name, '000000')
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 18000

    chart.set_categories(categories_ref)
    ws.add_chart(chart, anchor)

def make_chart_sheet(wb, sname, ds_title, raw_sname,
                     enc_col_start, dec_col_start, ratio_col_start,
                     y_label_spd, y_label_ratio):
    ws = wb.create_sheet(sname)
    ws.column_dimensions['A'].width = 8
    for col in range(2, NF + 2):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # ── Data table (formula refs) ─────────────────────────────────────────
    n_rows = len(ROUND_ORDER)
    # Header row 1: encode speed
    ws.cell(1, 1, 'Round').font = HDR_FONTW; ws.cell(1,1).fill = HDR_BLUE
    for i, fmt in enumerate(FORMATS):
        c = ws.cell(1, 2+i, f'enc-{fmt}'); c.font = HDR_FONT
    for i, fmt in enumerate(FORMATS):
        c = ws.cell(1, 2+NF+i, f'dec-{fmt}'); c.font = HDR_FONT
    for i, fmt in enumerate(FORMATS):
        c = ws.cell(1, 2+2*NF+i, f'ratio-{fmt}'); c.font = HDR_FONT

    # Data rows
    for idx, rnd in enumerate(ROUND_ORDER):
        raw_row = idx + 3   # raw sheet: header rows 1-2, data from row 3
        r = idx + 2         # chart sheet row (row 1 = header)
        fill = ERA_FILL[era_of(rnd)]
        ws.cell(r, 1, rnd).font = HDR_FONT; ws.cell(r, 1).fill = fill
        for i in range(NF):
            enc_col = get_column_letter(enc_col_start + i)
            dec_col = get_column_letter(dec_col_start + i)
            rat_col = get_column_letter(ratio_col_start + i)
            ws.cell(r, 2+i,       f"='{raw_sname}'!{enc_col}{raw_row}").fill = fill
            ws.cell(r, 2+NF+i,    f"='{raw_sname}'!{dec_col}{raw_row}").fill = fill
            ws.cell(r, 2+2*NF+i,  f"='{raw_sname}'!{rat_col}{raw_row}").fill = fill

    # ── Charts ────────────────────────────────────────────────────────────
    raw_ws = ws.parent[raw_sname]
    cat_ref = Reference(raw_ws, min_col=1, max_col=1, min_row=3, max_row=2+n_rows)

    # Encode Speed chart
    chart_enc = LineChart()
    chart_enc.title = f'{ds_title} — Encode Speed (MB/s)'
    chart_enc.style = 10
    chart_enc.y_axis.title = y_label_spd
    chart_enc.x_axis.title = 'Round'
    chart_enc.height = 14; chart_enc.width = 26
    def add_series(chart, raw_ws, col, fmt, n_rows):
        ref = Reference(raw_ws, min_col=col, max_col=col, min_row=3, max_row=2+n_rows)
        s = Series(ref, title=SeriesLabel(v=fmt))
        s.graphicalProperties.line.solidFill = SERIES_COLORS[fmt]
        s.graphicalProperties.line.width = 18000
        chart.series.append(s)

    for i, fmt in enumerate(FORMATS):
        add_series(chart_enc, raw_ws, enc_col_start+i, fmt, n_rows)
    chart_enc.set_categories(cat_ref)
    ws.add_chart(chart_enc, 'A' + str(n_rows + 5))

    # Decode Speed chart
    chart_dec = LineChart()
    chart_dec.title = f'{ds_title} — Decode Speed (MB/s)'
    chart_dec.style = 10
    chart_dec.y_axis.title = y_label_spd
    chart_dec.x_axis.title = 'Round'
    chart_dec.height = 14; chart_dec.width = 26
    for i, fmt in enumerate(FORMATS):
        add_series(chart_dec, raw_ws, dec_col_start+i, fmt, n_rows)
    chart_dec.set_categories(cat_ref)
    ws.add_chart(chart_dec, 'N' + str(n_rows + 5))

    # Compress Ratio chart
    chart_rat = LineChart()
    chart_rat.title = f'{ds_title} — Compress Ratio'
    chart_rat.style = 10
    chart_rat.y_axis.title = 'Ratio (lower = better)'
    chart_rat.y_axis.scaling.min = 0.75
    chart_rat.y_axis.scaling.max = 1.25
    chart_rat.x_axis.title = 'Round'
    chart_rat.height = 14; chart_rat.width = 26
    for i, fmt in enumerate(FORMATS):
        add_series(chart_rat, raw_ws, ratio_col_start+i, fmt, n_rows)
    chart_rat.set_categories(cat_ref)
    ws.add_chart(chart_rat, 'A' + str(n_rows + 35))

    return ws

# ── Build sheets ───────────────────────────────────────────────────────────
print("\nBuilding xlsx sheets...")
make_raw(wb, 'Raw_claw-code', 'claw-code')
make_raw(wb, 'Raw_llama_cpp', 'llama.cpp')
print("  Raw sheets done")

make_chart_sheet(wb, 'claw-code_Speed',
    'claw-code', 'Raw_claw-code',
    COL_ENC_SPD, COL_DEC_SPD, COL_RATIO,
    'MB/s', 'Ratio')
make_chart_sheet(wb, 'llama_cpp_Speed',
    'llama.cpp', 'Raw_llama_cpp',
    COL_ENC_SPD, COL_DEC_SPD, COL_RATIO,
    'MB/s', 'Ratio')
print("  Speed+Ratio chart sheets done")

# RSS chart sheet (encode only – decode unreliable)
def make_rss_sheet(wb, sname, ds_title, raw_sname):
    ws = wb.create_sheet(sname)
    n_rows = len(ROUND_ORDER)
    raw_ws = wb[raw_sname]
    ws.cell(1, 1, 'Round').font = HDR_FONTW; ws.cell(1,1).fill = HDR_BLUE
    for i, fmt in enumerate(FORMATS):
        ws.cell(1, 2+i, f'enc-{fmt}').font = HDR_FONT
    for idx, rnd in enumerate(ROUND_ORDER):
        raw_row = idx + 3; r = idx + 2
        fill = ERA_FILL[era_of(rnd)]
        ws.cell(r, 1, rnd).font = HDR_FONT; ws.cell(r, 1).fill = fill
        for i in range(NF):
            col_l = get_column_letter(COL_ENC_RSS + i)
            ws.cell(r, 2+i, f"='{raw_sname}'!{col_l}{raw_row}").fill = fill
    cat_ref = Reference(raw_ws, min_col=1, max_col=1, min_row=3, max_row=2+n_rows)
    chart = LineChart()
    chart.title = f'{ds_title} — Encode RSS (MB)'
    chart.style = 10
    chart.y_axis.title = 'RSS (MB)'; chart.x_axis.title = 'Round'
    chart.height = 14; chart.width = 26
    for i, fmt in enumerate(FORMATS):
        ref = Reference(raw_ws, min_col=COL_ENC_RSS+i, max_col=COL_ENC_RSS+i,
                        min_row=3, max_row=2+n_rows)
        s = Series(ref, title=SeriesLabel(v=fmt))
        s.graphicalProperties.line.solidFill = SERIES_COLORS[fmt]
        s.graphicalProperties.line.width = 18000
        chart.series.append(s)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, 'A' + str(n_rows + 5))
    return ws

make_rss_sheet(wb, 'claw-code_RSS', 'claw-code', 'Raw_claw-code')
make_rss_sheet(wb, 'llama_cpp_RSS', 'llama.cpp', 'Raw_llama_cpp')
print("  RSS chart sheets done")

# Energy chart sheet (encode only — decode not reliable)
def make_energy_sheet(wb, sname, ds_title, raw_sname):
    ws = wb.create_sheet(sname)
    n_rows = len(ROUND_ORDER)
    raw_ws = wb[raw_sname]
    ws.cell(1, 1, 'Round').font = HDR_FONTW; ws.cell(1,1).fill = HDR_BLUE
    for i, fmt in enumerate(FORMATS):
        ws.cell(1, 2+i, f'enc-{fmt}').font = HDR_FONT
    for idx, rnd in enumerate(ROUND_ORDER):
        raw_row = idx + 3; r = idx + 2
        fill = ERA_FILL[era_of(rnd)]
        ws.cell(r, 1, rnd).font = HDR_FONT; ws.cell(r, 1).fill = fill
        for i in range(NF):
            col_l = get_column_letter(COL_ENC_ENJ + i)
            ws.cell(r, 2+i, f"='{raw_sname}'!{col_l}{raw_row}").fill = fill
    cat_ref = Reference(raw_ws, min_col=1, max_col=1, min_row=3, max_row=2+n_rows)
    chart = LineChart()
    chart.title = f'{ds_title} — Encode CPU Energy (J)'
    chart.style = 10
    chart.y_axis.title = 'Energy (J)'; chart.x_axis.title = 'Round'
    chart.height = 14; chart.width = 26
    for i, fmt in enumerate(FORMATS):
        ref = Reference(raw_ws, min_col=COL_ENC_ENJ+i, max_col=COL_ENC_ENJ+i,
                        min_row=3, max_row=2+n_rows)
        s = Series(ref, title=SeriesLabel(v=fmt))
        s.graphicalProperties.line.solidFill = SERIES_COLORS[fmt]
        s.graphicalProperties.line.width = 18000
        chart.series.append(s)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, 'A' + str(n_rows + 5))
    return ws

make_energy_sheet(wb, 'claw-code_Energy', 'claw-code', 'Raw_claw-code')
make_energy_sheet(wb, 'llama_cpp_Energy', 'llama.cpp', 'Raw_llama_cpp')
print("  Energy chart sheets done")

OUT = str(SCRIPT_DIR / 'lzfse2_DOE_trend.xlsx')
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Rounds: {len(ROUND_ORDER)}, Formats: {FORMATS}")
